"""Susun rekap kehadiran dan tulis ke .xlsx mengikuti format 'data jadi'."""

import io
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .sesi import label_tanggal

TIPIS = Side(style="thin")
TEBAL = Side(style="medium")
KOTAK_TIPIS = Border(left=TIPIS, right=TIPIS, top=TIPIS, bottom=TIPIS)
KOTAK_TEBAL = Border(left=TEBAL, right=TEBAL, top=TEBAL, bottom=TEBAL)
TENGAH = Alignment(horizontal="center", vertical="center", wrap_text=True)
KIRI = Alignment(horizontal="left", vertical="center", wrap_text=True)
ARSIR_HEADER = PatternFill("solid", fgColor="DDEBF7")
ARSIR_ALPA = PatternFill("solid", fgColor="FCE4E4")

KOLOM_MULAI_SESI = 4  # kolom D
KOLOM_PER_SESI = 4  # Status | Ceklog 1 | Ceklog 2 | Durasi (jam)


def _ke_menit(teks):
    """'08.50' / '08:50' / '0850' -> menit sejak tengah malam."""
    m = re.match(r"^\s*(\d{1,2})\s*[.:]?\s*(\d{2})\s*$", str(teks))
    if not m:
        raise ValueError(f"Format jam tidak valid: {teks!r}")
    return int(m.group(1)) * 60 + int(m.group(2))


def _dari_menit(menit):
    return f"{int(menit) // 60:02d}:{int(menit) % 60:02d}"


def susun_rekap(log, sesi, roster, toleransi_awal=15, toleransi_akhir=15, override=None):
    """Hitung status kehadiran tiap peserta pada tiap sesi.

    log       : DataFrame uid, nama, tanggal, jam, menit
    sesi      : list dict {tanggal, nama, jam_mulai, jam_selesai}
    roster    : DataFrame uid, nim, nama  (urutan baris = urutan di laporan)
    override  : DataFrame uid, tanggal, sesi_ke, status  (untuk S / I)

    Hasil: list baris {no, nim, nama, sel: [(status, waktu), ...]}
    """
    override_map = {}
    if override is not None and not override.empty:
        for r in override.itertuples(index=False):
            override_map[(int(r.uid), int(r.sesi_ke))] = str(r.status).strip().upper()[:1]

    jendela = []
    for s in sesi:
        tgl = pd.Timestamp(s["tanggal"]).normalize()
        mulai = _ke_menit(s["jam_mulai"]) - toleransi_awal
        selesai = _ke_menit(s["jam_selesai"]) + toleransi_akhir
        jendela.append((tgl, mulai, selesai))

    # indeks cepat: (uid, tanggal) -> daftar menit
    indeks = {}
    for r in log.itertuples(index=False):
        indeks.setdefault((int(r.uid), pd.Timestamp(r.tanggal).normalize()), []).append(int(r.menit))

    baris = []
    for i, p in enumerate(roster.itertuples(index=False), 1):
        uid = int(p.uid)
        sel = []
        for k, (tgl, mulai, selesai) in enumerate(jendela, 1):
            manual = override_map.get((uid, k))
            if manual in ("S", "I", "A", "H"):
                sel.append((manual, "MANUAL", "MANUAL", None))
                continue

            cocok = sorted(m for m in indeks.get((uid, tgl), []) if mulai <= m <= selesai)
            if not cocok:
                sel.append(("A", "-", "-", None))
                continue

            ceklog1 = cocok[0]
            ceklog2 = cocok[1] if len(cocok) > 1 else None
            durasi = round((ceklog2 - ceklog1) / 60, 2) if ceklog2 is not None else None
            sel.append(
                (
                    "H",
                    _dari_menit(ceklog1),
                    _dari_menit(ceklog2) if ceklog2 is not None else "-",
                    durasi,
                )
            )
        total_jam = round(sum(s[3] for s in sel if s[3] is not None), 2)
        baris.append(
            {
                "no": i,
                "nim": str(p.nim),
                "nama": str(p.nama),
                "sel": sel,
                "total_jam": total_jam,
            }
        )
    return baris


def tulis_xlsx(baris, sesi, judul, meta, nama_sheet="Laporan_Absen"):
    """Tulis rekap ke .xlsx dengan tata letak sama seperti file referensi."""
    n_sesi = len(sesi)
    wb = Workbook()
    ws = wb.active
    ws.title = nama_sheet[:31]

    kol_rekap = KOLOM_MULAI_SESI + KOLOM_PER_SESI * n_sesi
    kol_a, kol_s, kol_i = kol_rekap + 1, kol_rekap + 2, kol_rekap + 3
    kol_total, kol_jam, kol_persen = kol_rekap + 4, kol_rekap + 5, kol_rekap + 6
    kol_akhir = kol_persen

    # --- judul ---
    for r, teks in enumerate(judul, 1):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=kol_akhir)
        c = ws.cell(r, 1, teks)
        c.font = Font(bold=True, size=13.5)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # --- keterangan blok / kelas ---
    baris_meta = len(judul) + 2
    for i, teks in enumerate(meta):
        c = ws.cell(baris_meta + i, 1, teks)
        c.font = Font(bold=True, size=12)

    # --- header tabel (4 tingkat) ---
    h0 = baris_meta + len(meta) + 1
    h1, h2, h3 = h0 + 1, h0 + 2, h0 + 3

    for kol, judul_kol in ((1, "No."), (2, "NIM"), (3, "Nama")):
        ws.merge_cells(start_row=h0, start_column=kol, end_row=h3, end_column=kol)
        c = ws.cell(h0, kol, judul_kol)
        c.font = Font(bold=True, size=9)

    for k, s in enumerate(sesi):
        c1 = KOLOM_MULAI_SESI + KOLOM_PER_SESI * k
        c2 = c1 + KOLOM_PER_SESI - 1
        for r, teks in (
            (h0, label_tanggal(s["tanggal"])),
            (h1, s.get("nama") or f"SESI {k + 1}"),
            (h2, f"{s['jam_mulai']} - {s['jam_selesai']}"),
        ):
            ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
            ws.cell(r, c1, teks).font = Font(bold=True, size=8)
        for off, teks in enumerate(("Status", "Ceklog 1", "Ceklog 2", "Durasi\n(jam)")):
            ws.cell(h3, c1 + off, teks).font = Font(bold=True, size=8)

    ws.merge_cells(start_row=h0, start_column=kol_rekap, end_row=h2, end_column=kol_i)
    ws.cell(h0, kol_rekap, "Jumlah Kehadiran").font = Font(bold=True, size=8)
    for kol, huruf in ((kol_rekap, "H"), (kol_a, "A"), (kol_s, "S"), (kol_i, "I")):
        ws.cell(h3, kol, huruf).font = Font(bold=True, size=8)

    for kol, teks in (
        (kol_total, "Total Kehadiran"),
        (kol_jam, "Total Jam"),
        (kol_persen, "Persentasi (%)"),
    ):
        ws.merge_cells(start_row=h0, start_column=kol, end_row=h3, end_column=kol)
        ws.cell(h0, kol, teks).font = Font(bold=True, size=8)

    for r in range(h0, h3 + 1):
        for c in range(1, kol_akhir + 1):
            sel = ws.cell(r, c)
            sel.alignment = TENGAH
            sel.border = KOTAK_TEBAL
            sel.fill = ARSIR_HEADER

    # --- isi data ---
    r0 = h3 + 1
    for i, b in enumerate(baris):
        r = r0 + i
        ws.cell(r, 1, b["no"]).alignment = TENGAH
        ws.cell(r, 2, b["nim"]).alignment = TENGAH
        ws.cell(r, 3, b["nama"]).alignment = KIRI
        L = get_column_letter
        sel_durasi = []
        for k, (status, ceklog1, ceklog2, durasi) in enumerate(b["sel"]):
            c1 = KOLOM_MULAI_SESI + KOLOM_PER_SESI * k
            sc = ws.cell(r, c1, status)
            sc.font = Font(bold=True, size=8)
            sc.alignment = TENGAH
            if status == "A":
                sc.fill = ARSIR_ALPA
            ws.cell(r, c1 + 1, ceklog1).alignment = TENGAH
            ws.cell(r, c1 + 2, ceklog2).alignment = TENGAH
            dc = ws.cell(r, c1 + 3, durasi)
            dc.alignment = TENGAH
            dc.number_format = "0.00"
            sel_durasi.append(f"{L(c1 + 3)}{r}")

        rentang = f"{L(KOLOM_MULAI_SESI)}{r}:{L(kol_rekap - 1)}{r}"
        ws.cell(r, kol_rekap, f'=COUNTIF({rentang},"H")')
        ws.cell(r, kol_a, f'=COUNTIF({rentang},"A")')
        ws.cell(r, kol_s, f'=COUNTIF({rentang},"S")')
        ws.cell(r, kol_i, f'=COUNTIF({rentang},"I")')
        ws.cell(r, kol_total, f"={L(kol_rekap)}{r}+{L(kol_s)}{r}+{L(kol_i)}{r}")

        jam = ws.cell(r, kol_jam, f"=SUM({','.join(sel_durasi)})" if sel_durasi else 0)
        jam.number_format = "0.00"
        persen = ws.cell(r, kol_persen, f"={L(kol_total)}{r}/{n_sesi}")
        persen.number_format = "0%"

        for c in range(1, kol_akhir + 1):
            sel = ws.cell(r, c)
            sel.border = KOTAK_TIPIS
            if sel.font.size != 8:
                sel.font = Font(size=8)
            if c >= KOLOM_MULAI_SESI:
                sel.alignment = TENGAH

    # --- lebar kolom ---
    ws.column_dimensions["A"].width = 4.5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 28
    for k in range(n_sesi):
        c1 = KOLOM_MULAI_SESI + KOLOM_PER_SESI * k
        for off, lebar in enumerate((6, 8, 8, 7)):
            ws.column_dimensions[get_column_letter(c1 + off)].width = lebar
    for kol in (kol_rekap, kol_a, kol_s, kol_i):
        ws.column_dimensions[get_column_letter(kol)].width = 4.5
    ws.column_dimensions[get_column_letter(kol_total)].width = 9
    ws.column_dimensions[get_column_letter(kol_jam)].width = 9
    ws.column_dimensions[get_column_letter(kol_persen)].width = 10
    ws.freeze_panes = ws.cell(r0, KOLOM_MULAI_SESI)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
